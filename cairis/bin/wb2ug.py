#!/usr/bin/python3
#  Licensed to the Apache Software Foundation (ASF) under one
#  or more contributor license agreements.  See the NOTICE file
#  distributed with this work for additional information
#  regarding copyright ownership.  The ASF licenses this file
#  to you under the Apache License, Version 2.0 (the
#  "License"); you may not use this file except in compliance
#  with the License.  You may obtain a copy of the License at
#
#  http://www.apache.org/licenses/LICENSE-2.0
#
#  Unless required by applicable law or agreed to in writing,
#  software distributed under the License is distributed on an
#  "AS IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY
#  KIND, either express or implied.  See the License for the
#  specific language governing permissions and limitations
#  under the License.

import argparse
import sys
from openpyxl import load_workbook


__author__ = 'Shamal Faily'

def main(args=None):
  parser = argparse.ArgumentParser(description='Computer Aided Integration of Requirements and Information Security - Workbook to XML converter')
  parser.add_argument('modelFile',help='CAIRIS model file output')
  parser.add_argument('--xlsx',dest='xlsxFile',help='Workbook to input', default='')
  args = parser.parse_args()

  if (args.xlsxFile == ''):
    raise Exception('Workbook file not specified')

  xmlBuf = '<?xml version="1.0"?>\n<!DOCTYPE cairis_model PUBLIC "-//CAIRIS//DTD MODEL 1.0//EN" "https://cairis.org/dtd/cairis_model.dtd">\n\n<cairis_model>\n<synopses>\n'

  wb = load_workbook(filename = args.xlsxFile,data_only=True)
  ugSheet = wb.worksheets[0]
  for row in ugSheet.iter_rows(min_row=2):
    refName = row[0].value
    refDesc = row[1].value
    pName = row[2].value
    pdRef = row[3].value
    elType = row[4].value
    ugName = row[5].value
    if (ugName != '' and ugName != None and ugName != 0):
      initSat = row[6].value
      if (pdRef == 'persona'):
        xmlBuf += '  <characteristic_synopsis characteristic="' + refName + '" synopsis="' + ugName + '" dimension="' + elType + '" actor_type="persona" actor="' + pName + '" satisfaction="' + initSat + '" />\n'
      else:
        xmlBuf += '  <reference_synopsis reference="' + refName + '" synopsis="' + ugName + '" dimension="' + elType + '" actor_type="persona" actor="' + pName + '" satisfaction="' + initSat + '" />\n'
      

  contSheet = wb.worksheets[1]
  for row in contSheet.iter_rows(min_row=2):
    srcName = row[0].value
    destName = row[1].value
    meName = row[2].value
    contName = row[3].value
    if (srcName != '' and srcName != None and srcName != 0 and destName != '' and destName != None and destName != 0):
      xmlBuf += '  <reference_contribution source="' + srcName + '" destination="' + destName + '" means_end="' + meName + '" contribution="' + contName + '" />\n'
  xmlBuf += "</synopses>\n</cairis_model>"

  f = open(args.modelFile,'w')
  f.write(xmlBuf)
  f.close()

if __name__ == '__main__':
  try:
    main()
  except Exception as e:
    print('Fatal wb2ug error: ' + str(e))
    sys.exit(-1)
