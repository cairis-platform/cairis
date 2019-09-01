#!/usr/bin/python3
#  Licensed to the Apache Software Foundation (ASF) under one
#  or more contributor license agreements.  See the NOTICE file
#  distributed with this work for additional information
#  regarding copyright ownership.  The ASF licenses this file
#  to you under the Apache License, Version 2.0 (the
#  "License"); you may not use this file except in compliance
#  with the License.  You may obtain a copy of the License at
#
#  http://www.apache.org/licenses/LICENSE-2.0
#
#  Unless required by applicable law or agreed to in writing,
#  software distributed under the License is distributed on an
#  "AS IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY
#  KIND, either express or implied.  See the License for the
#  specific language governing permissions and limitations
#  under the License.

import argparse
import os
from cairis.core.ARM import *
from openpyxl import load_workbook

__author__ = 'Shamal Faily'

def main(args=None):
  parser = argparse.ArgumentParser(description='Computer Aided Integration of Requirements and Information Security - Requirements spreadsheet to XML')
  parser.add_argument('modelFile',help='model file to create')
  parser.add_argument('--spreadsheet',dest='ssName',help='Spreadsheet file (.xlsx)')
  args = parser.parse_args()
  if not os.path.exists(args.ssName):
    raise ARMException(args.ssName + ' does not exist')
  convertRequirementSpreadsheet(args.ssName,args.modelFile)

def convertRequirementSpreadsheet(ssName,modelFile):
  wb = load_workbook(ssName)
  if (len(wb.sheetnames) > 1):
    raise ARMException('No more than one sheet expected')
  ws = wb.active

  expectedHdrCols = set(['Requirement','Description','Priority','Environment/Asset','Reference','Rationale','Fit Criterion','Originator','Type'])
  cellDict = {}
  actualHdrCols = set()
  for col in ws.iter_cols(min_row=1, max_row=1):
    for cell in col:
      if cell.value in expectedHdrCols:
        cellDict[cell.col_idx] = cell.value
      actualHdrCols.add(cell.value)

  if (actualHdrCols == expectedHdrCols) or (expectedHdrCols < actualHdrCols):
    xmlBuf = '<?xml version="1.0"?>\n<!DOCTYPE goals PUBLIC "-//CAIRIS//DTD GOALS USABILITY 1.0//EN" "http://cairis.org/dtd/goals.dtd">\n\n<goals>\n\n'
    reqLabel = 1
    for row in ws.iter_rows(min_row=2):
      rowDict = {}
      for cell in row:
        reqAttribute = cellDict[cell.col_idx]
        rowDict[reqAttribute] = cell.value
      xmlBuf += '<requirement name=\"' + rowDict['Requirement'] + '\" reference=\"' + rowDict['Reference'] + '\" reference_type=\"' + rowDict['Environment/Asset'].lower() + '\" label=\"' + str(reqLabel) + '\" type=\"' + rowDict['Type'] + '\" priority=\"' + str(rowDict['Priority']) + '\">\n  <description>' + rowDict['Description'] + '</description>\n  <rationale>' + rowDict['Rationale'] + '</rationale>\n  <fit_criterion>' + rowDict['Fit Criterion'] + '</fit_criterion>\n  <originator>' + rowDict['Originator'] + '</originator>\n</requirement>\n'
      reqLabel += 1
    xmlBuf += '</goals>'
    f = open(modelFile,'w')
    f.write(xmlBuf)
    f.close()
  else:
    raise ARMException("Spreadsheet headings don't match expecting headings") 

if __name__ == '__main__':
  main()
