#  Licensed to the Apache Software Foundation (ASF) under one
#  or more contributor license agreements.  See the NOTICE file
#  distributed with this work for additional information
#  regarding copyright ownership.  The ASF licenses this file
#  to you under the Apache License, Version 2.0 (the
#  "License"); you may not use this file except in compliance
#  with the License.  You may obtain a copy of the License at
#
#  http://www.apache.org/licenses/LICENSE-2.0
#
#  Unless required by applicable law or agreed to in writing,
#  software distributed under the License is distributed on an
#  "AS IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY
#  KIND, either express or implied.  See the License for the
#  specific language governing permissions and limitations
#  under the License.


from .SecurityPatternContentHandler import SecurityPatternContentHandler
from .AttackPatternContentHandler import AttackPatternContentHandler
from .TVTypeContentHandler import TVTypeContentHandler
from .DomainValueContentHandler import DomainValueContentHandler
from .DirectoryContentHandler import DirectoryContentHandler
from .RiskAnalysisContentHandler import RiskAnalysisContentHandler
from .GoalsContentHandler import GoalsContentHandler
from .UsabilityContentHandler import UsabilityContentHandler
from .MisusabilityContentHandler import MisusabilityContentHandler
from .AssociationsContentHandler import AssociationsContentHandler
from .CairisContentHandler import CairisContentHandler
from .ArchitecturalPatternContentHandler import ArchitecturalPatternContentHandler
from .SynopsesContentHandler import SynopsesContentHandler
from .StoriesContentHandler import StoriesContentHandler
from .TemplateAssetsContentHandler import TemplateAssetsContentHandler
from .ProcessesContentHandler import ProcessesContentHandler
from .LocationsContentHandler import LocationsContentHandler
from .DataflowsContentHandler import DataflowsContentHandler
from .DiagramsNetContentHandler import DiagramsNetContentHandler
from cairis.core.DataFlowParameters import DataFlowParameters
from cairis.core.AssetParameters import AssetParameters
from cairis.core.AssetEnvironmentProperties import AssetEnvironmentProperties
from cairis.core.ClassAssociationParameters import ClassAssociationParameters
from cairis.core.TrustBoundary import TrustBoundary
from cairis.core.ReferenceSynopsis import ReferenceSynopsis
from cairis.core.ReferenceContribution import ReferenceContribution
from cairis.core.ExternalDocumentParameters import ExternalDocumentParameters
from cairis.core.DocumentReferenceParameters import DocumentReferenceParameters
from cairis.core.PersonaCharacteristicParameters import PersonaCharacteristicParameters
from cairis.core.Borg import Borg
import cairis.core.DefaultParametersFactory
import xml.sax
from openpyxl import load_workbook
from cairis.core.ARM import *

__author__ = 'Shamal Faily'

def importSecurityPatternsFile(importFile,session_id = None):
  try:
    parser = xml.sax.make_parser()
    handler = SecurityPatternContentHandler()
    parser.setContentHandler(handler)
    parser.setEntityResolver(handler)
    parser.parse(importFile)
    taps = handler.assets()
    spps = handler.patterns()
    vts = handler.metricTypes()
    return importSecurityPatterns(taps,spps,vts,session_id)
  except xml.sax.SAXException as e:
    raise ARMException("Error parsing" + importFile + ": " + e.getMessage())

def importSecurityPatterns(taps,spps,vts,session_id=None):
  noOfTaps = len(taps)
  noOfSpps = len(spps)

  b = Borg()
  db_proxy = b.get_dbproxy(session_id)

  for tap in taps:
    if (db_proxy.nameExists(tap.name(),'asset')):
      raise ARMException("Cannot import template asset with name " + tap.name() + " as asset with same name already exists.")
  db_proxy.addSecurityPatterns(vts,taps,spps)
  msgStr =  'Imported ' + str(noOfTaps) + ' template assets and ' + str(noOfSpps) + ' security patterns'
  return msgStr

def importAttackPattern(importFile,session_id = None):
  try:
    parser = xml.sax.make_parser()
    handler = AttackPatternContentHandler(session_id = session_id)
    parser.setContentHandler(handler)
    parser.setEntityResolver(handler)
    parser.parse(importFile)
    assets = handler.assets()
    attackers = handler.attackers()
    vulnerability = handler.vulnerability()
    threat = handler.threat()
    risk = handler.risk()
    raTxt = importRiskAnalysis([],assets,[vulnerability],attackers,[threat],[risk],[],[],session_id)
    obsTxt = importRequirements([],[],handler.obstacles(),[],[],[],session_id)
    assocTxt = importAssociations([],handler.obstacleAssociations(),[],session_id)
    return obsTxt + assocTxt + raTxt
  except xml.sax.SAXException as e:
    raise ARMException("Error parsing" + importFile + ": " + e.getMessage())

def importTVTypeFile(importFile,isOverwrite=1,session_id = None):
  try:
    parser = xml.sax.make_parser()
    handler = TVTypeContentHandler()
    parser.setContentHandler(handler)
    parser.setEntityResolver(handler)
    parser.parse(importFile)
    vulTypes,threatTypes = handler.types()
    return importTVTypes(vulTypes,threatTypes,isOverwrite,session_id)
  except xml.sax.SAXException as e:
    raise ARMException("Error parsing" + importFile + ": " + e.getMessage())
 
def importTVTypes(vulTypes,threatTypes,isOverwrite,session_id):
  b = Borg()
  db_proxy = b.get_dbproxy(session_id)
  noOfVts = len(vulTypes)
  noOfTts = len(threatTypes)
  if (noOfVts > 0):
    if (isOverwrite):
      db_proxy.deleteVulnerabilityType(-1)
    for vt in vulTypes:
      db_proxy.addValueType(vt)
  if (noOfTts > 0):
    if (isOverwrite):
      db_proxy.deleteThreatType(-1)
    for tt in threatTypes:
      db_proxy.addValueType(tt)
  msgStr = 'Imported ' + str(noOfVts) + ' vulnerability types and ' + str(noOfTts) + ' threat types.'
  return msgStr

def importDirectoryFile(importFile,isOverwrite=1,session_id = None):
  try:
    parser = xml.sax.make_parser()
    handler = DirectoryContentHandler()
    parser.setContentHandler(handler)
    parser.setEntityResolver(handler)
    parser.parse(importFile)
    vulDir,threatDir = handler.directories()
    vdSize = len(vulDir)
    tdSize = len(threatDir)
    b = Borg()
    db_proxy = b.get_dbproxy(session_id)
    if (vdSize > 0):
      db_proxy.addVulnerabilityDirectory(vulDir,isOverwrite)
    if (tdSize > 0):
      db_proxy.addThreatDirectory(threatDir,isOverwrite)
    msgStr = 'Imported ' + str(vdSize) + ' template vulnerabilities and ' + str(tdSize) + ' template threats.'
    return msgStr
  except xml.sax.SAXException as e:
    raise ARMException("Error parsing" + importFile + ": " + e.getMessage())


def importRequirementsFile(importFile,session_id = None):
  try:
    parser = xml.sax.make_parser()
    handler = GoalsContentHandler(session_id = session_id)
    parser.setContentHandler(handler)
    parser.setEntityResolver(handler)
    parser.parse(importFile)
    return importRequirements(handler.domainProperties(),handler.goals(),handler.obstacles(),handler.requirements(),handler.usecases(),handler.countermeasures(),session_id = session_id)
  except xml.sax.SAXException as e:
    raise ARMException("Error parsing" + importFile + ": " + e.getMessage())

def importRequirementsString(buf,session_id = None):
  try:
    handler = GoalsContentHandler(session_id = session_id)
    xml.sax.parseString(buf,handler)
    return importRequirements(handler.domainProperties(),handler.goals(),handler.obstacles(),handler.requirements(),handler.usecases(),handler.countermeasures(),session_id = session_id)
  except xml.sax.SAXException as e:
    raise ARMException("Error parsing imported file: " + e.getMessage())

def importRequirements(dpParameterSet,goalParameterSet,obsParameterSet,reqParameterSet,ucParameterSet,cmParameterSet,session_id):
  b = Borg()
  db_proxy = b.get_dbproxy(session_id)
  dpCount = 0
  for dpParameters in dpParameterSet:
    objtId = db_proxy.existingObject(dpParameters.name(),'domainproperty')
    if objtId == -1:
      db_proxy.addDomainProperty(dpParameters)
    else:
      dpParameters.setId(objtId)
      db_proxy.updateDomainProperty(dpParameters)
    dpCount += 1

  goalCount = 0
  for goalParameters in goalParameterSet:
    objtId = db_proxy.existingObject(goalParameters.name(),'goal')
    if objtId == -1:
      db_proxy.addGoal(goalParameters)
    else:
      goalParameters.setId(objtId)
      db_proxy.updateGoal(goalParameters)
    goalCount += 1

  obsCount = 0
  for obsParameters in obsParameterSet:
    objtId = db_proxy.existingObject(obsParameters.name(),'obstacle')
    if objtId == -1:
      db_proxy.addObstacle(obsParameters)
    else:
      obsParameters.setId(objtId)
      db_proxy.updateObstacle(obsParameters)
    obsCount += 1

  reqCount = 0
  for req in reqParameterSet:
    objtId = db_proxy.existingObject(req.name(),'requirement')
    if objtId == -1:
      isAsset = True
      if (req.domainType() == 'environment'):
        isAsset = False
      db_proxy.addRequirement(req,req.domain(),isAsset)
    else:
      db_proxy.updateRequirement(req)
    reqCount += 1

  ucCount = 0
  for ucParameters in ucParameterSet:
    objtId = db_proxy.existingObject(ucParameters.name(),'usecase')
    if objtId == -1:
      db_proxy.addUseCase(ucParameters)
    else:
      ucParameters.setId(objtId)
      db_proxy.updateUseCase(ucParameters)
    ucCount += 1

  cmCount = 0
  for cmParameters in cmParameterSet:
    objtId = db_proxy.existingObject(cmParameters.name(),'countermeasure')
    if objtId == -1:
      db_proxy.addCountermeasure(cmParameters)
    else:
      cmParameters.setId(objtId)
      db_proxy.updateCountermeasure(cmParameters)
    cmCount += 1
  msgStr = 'Imported ' + str(dpCount) + ' domain properties, ' + str(goalCount) + ' goals, ' + str(obsCount) + ' obstacles, ' + str(reqCount) + ' requirements, ' + str(ucCount) + ' use cases, and ' + str(cmCount) + ' countermeasures.'
  return msgStr

def importRiskAnalysisFile(importFile,session_id = None):
  try:
    parser = xml.sax.make_parser()
    handler = RiskAnalysisContentHandler()
    parser.setContentHandler(handler)
    parser.setEntityResolver(handler)
    parser.parse(importFile)
    return importRiskAnalysis(handler.roles(),handler.assets(),handler.vulnerabilities(),handler.attackers(),handler.threats(),handler.risks(),handler.responses(),handler.associations(),session_id = session_id)
  except xml.sax.SAXException as e:
    raise ARMException("Error parsing" + importFile + ": " + e.getMessage())

def importRiskAnalysis(roleParameterSet,assetParameterSet,vulParameterSet,attackerParameterSet,threatParameterSet,riskParameterSet,responseParameterSet,assocParameterSet,session_id):

  b = Borg()
  db_proxy = b.get_dbproxy(session_id)
  roleCount = 0
  for roleParameters in roleParameterSet:
    objtId = db_proxy.existingObject(roleParameters.name(),'role')
    if objtId == -1:
      db_proxy.addRole(roleParameters)
    else:
      roleParameters.setId(objtId)
      db_proxy.updateRole(roleParameters)
    roleCount += 1

  assetCount = 0
  for assetParameters in assetParameterSet:
    objtId = db_proxy.existingObject(assetParameters.name(),'asset')
    if objtId == -1:
      db_proxy.addAsset(assetParameters)
    else:
      assetParameters.setId(objtId)
      db_proxy.updateAsset(assetParameters)
    assetCount += 1

  vulCount = 0
  for vulParameters in vulParameterSet:
    objtId = db_proxy.existingObject(vulParameters.name(),'vulnerability')
    if objtId == -1:
      db_proxy.addVulnerability(vulParameters)
    else:
      vulParameters.setId(objtId)
      db_proxy.updateVulnerability(vulParameters)
    vulCount += 1

  attackerCount = 0
  for attackerParameters in attackerParameterSet:
    objtId = db_proxy.existingObject(attackerParameters.name(),'attacker')
    if objtId == -1:
      db_proxy.addAttacker(attackerParameters)
    else:
      attackerParameters.setId(objtId)
      db_proxy.updateAttacker(attackerParameters)
    attackerCount += 1

  threatCount = 0
  for threatParameters in threatParameterSet:
    objtId = db_proxy.existingObject(threatParameters.name(),'threat')
    if objtId == -1:
      db_proxy.addThreat(threatParameters)
    else:
      threatParameters.setId(objtId)
      db_proxy.updateThreat(threatParameters)
    threatCount += 1

  riskCount = 0
  for riskParameters in riskParameterSet:
    objtId = db_proxy.existingObject(riskParameters.name(),'risk')
    if objtId == -1:
      db_proxy.addRisk(riskParameters)
    else:
      riskParameters.setId(objtId)
      db_proxy.updateRisk(riskParameters)
    riskCount += 1

  responseCount = 0
  for responseParameters in responseParameterSet:
    objtId = db_proxy.existingObject(responseParameters.name(),'response')
    if objtId == -1:
      db_proxy.addResponse(responseParameters)
    else:
      responseParameters.setId(objtId)
      db_proxy.updateResponse(responseParameters)
    responseCount += 1

  rshipCount = 0
  for assocParameters in assocParameterSet:
    db_proxy.addClassAssociation(assocParameters)
    rshipCount += 1

  msgStr = 'Imported ' + str(roleCount) + ' roles, ' + str(assetCount) + ' assets, ' + str(vulCount) + ' vulnerabilities, ' + str(attackerCount) + ' attackers, ' + str(threatCount) + ' threats, ' + str(riskCount) + ' risks, ' + str(responseCount) + ' responses, and ' + str(rshipCount) + ' asset associations.'
  return msgStr

def importUsabilityFile(importFile,session_id = None):
  try:
    parser = xml.sax.make_parser()
    handler = UsabilityContentHandler()
    parser.setContentHandler(handler)
    parser.setEntityResolver(handler)
    parser.parse(importFile)
    return importUsability(handler.personas(),handler.externalDocuments(),handler.documentReferences(),handler.personaCharacteristics(),handler.tasks(),session_id=session_id)
  except xml.sax.SAXException as e:
    raise ARMException("Error parsing" + importFile + ": " + e.getMessage())


def importUsability(personaParameterSet,edParameterSet,drParameterSet,pcParameterSet,taskParameterSet,session_id):
  b = Borg()
  db_proxy = b.get_dbproxy(session_id)
  personaCount = 0
  for personaParameters in personaParameterSet:
    objtId = db_proxy.existingObject(personaParameters.name(),'persona')
    if objtId == -1:
      db_proxy.addPersona(personaParameters)
    else:
      personaParameters.setId(objtId)
      db_proxy.updatePersona(personaParameters)
    personaCount += 1

  edCount = 0
  for edParameters in edParameterSet:
    objtId = db_proxy.existingObject(edParameters.name(),'external_document')
    if objtId == -1:
      db_proxy.addExternalDocument(edParameters)
    else:
      edParameters.setId(objtId)
      db_proxy.updateExternalDocument(edParameters)
    edCount += 1

  drCount = 0
  for drParameters in drParameterSet:
    objtId = db_proxy.existingObject(drParameters.name(),'document_reference')
    if objtId == -1:
      db_proxy.addDocumentReference(drParameters)
    else:
      drParameters.setId(objtId)
      db_proxy.updateDocumentReference(drParameters)
    drCount += 1

  taskCount = 0
  for taskParameters in taskParameterSet:
    objtId = db_proxy.existingObject(taskParameters.name(),'task')
    if objtId == -1:
      db_proxy.addTask(taskParameters)
    else:
      taskParameters.setId(objtId)
      db_proxy.updateTask(taskParameters)
    taskCount += 1


  pcCount = 0
  for pcParameters in pcParameterSet:
    db_proxy.addPersonaCharacteristic(pcParameters)
    pcCount += 1

  msgStr = 'Imported ' + str(personaCount) + ' personas, ' + str(edCount) + ' external documents, ' + str(drCount) + ' document references, ' + str(pcCount) + ' persona characteristics, ' + str(taskCount) + ' tasks.'
  return msgStr


def importMisusabilityFile(importFile,session_id = None):
  try:
    parser = xml.sax.make_parser()
    handler = MisusabilityContentHandler()
    parser.setContentHandler(handler)
    parser.setEntityResolver(handler)
    parser.parse(importFile)
    return importMisusability(handler.conceptReferences(),handler.taskCharacteristics(),session_id=session_id)
  except xml.sax.SAXException as e:
    raise ARMException("Error parsing" + importFile + ": " + e.getMessage())

def importMisusability(crParameterSet,tcParameterSet,session_id):
  b = Borg()
  db_proxy = b.get_dbproxy(session_id)
  crCount = 0
  for crParameters in crParameterSet:
    objtId = db_proxy.existingObject(crParameters.name(),'concept_reference')
    if objtId == -1:
      db_proxy.addConceptReference(crParameters)
    else:
      crParameters.setId(objtId)
      db_proxy.updateConceptReference(crParameters)
    crCount += 1

  tcCount = 0
  for tcParameters in tcParameterSet:
    objtId = db_proxy.existingObject(tcParameters.task(),'task_characteristic')
    if objtId == -1:
      db_proxy.addTaskCharacteristic(tcParameters)
    else:
      tcParameters.setId(objtId)
      db_proxy.updateTaskCharacterisric(tcParameters)
    tcCount += 1
  msgStr = 'Imported ' + str(crCount) + ' concept references, and ' + str(tcCount) + ' task characteristics.'
  return msgStr


def importAssociationsFile(importFile,session_id = None):
  try:
    parser = xml.sax.make_parser()
    handler = AssociationsContentHandler(session_id = session_id)
    parser.setContentHandler(handler)
    parser.setEntityResolver(handler)
    parser.parse(importFile)
    return importAssociations(handler.manualAssociations(),handler.goalAssociations(),handler.dependencyAssociations(),session_id = session_id)
  except xml.sax.SAXException as e:
    raise ARMException("Error parsing" + importFile + ": " + e.getMessage())

def importAssociationsString(buf,session_id = None):
  try:
    handler = AssociationsContentHandler(session_id = session_id)
    xml.sax.parseString(buf,handler)
    return importAssociations(handler.manualAssociations(),handler.goalAssociations(),handler.dependencyAssociations(),session_id = session_id)
  except xml.sax.SAXException as e:
    raise ARMException("Error parsing" + importFile + ": " + e.getMessage())

  
def importAssociations(maParameterSet,gaParameterSet,depParameterSet,session_id):
  b = Borg()
  db_proxy = b.get_dbproxy(session_id)
  maCount = 0
  for tTable,fromId,toId,refType in maParameterSet:
    db_proxy.addTrace(tTable,fromId,toId,refType)
    maCount += 1
  gaCount = 0
  for gaParameters in gaParameterSet:
    db_proxy.addGoalAssociation(gaParameters)
    gaCount += 1
  depCount = 0
  for depParameters in depParameterSet:
    db_proxy.addDependency(depParameters)
    depCount += 1
  msgStr = 'Imported ' + str(maCount) + ' manual associations, ' + str(gaCount) + ' goal associations, and ' + str(depCount) + ' dependency associations.'
  return msgStr

def importProjectFile(importFile,session_id = None):
  try:
    parser = xml.sax.make_parser()
    handler = CairisContentHandler()
    parser.setContentHandler(handler)
    parser.setEntityResolver(handler)
    parser.parse(importFile)
    pSettings = handler.settings()
    envParameterSet = handler.environments()
    ceParameterSet = handler.compositeEnvironments()
    return importProjectData(pSettings,envParameterSet,ceParameterSet,session_id = session_id)
  except xml.sax.SAXException as e:
    raise ARMException("Error parsing" + importFile + ": " + e.getMessage())

def importProjectData(pSettings,envParameterSet,ceParameterSet,session_id):
  b = Borg()
  db_proxy = b.get_dbproxy(session_id)
  if (pSettings != None):
    db_proxy.updateSettings(pSettings[0],pSettings[1],pSettings[2],pSettings[3],pSettings[4],pSettings[5],pSettings[6],pSettings[7])
  envCount = 0
  for envParameters in (envParameterSet + ceParameterSet):
    objtId = db_proxy.existingObject(envParameters.name(),'environment')
    if objtId == -1:
      db_proxy.addEnvironment(envParameters)
    else:
      envParameters.setId(objtId)
      db_proxy.updateEnvironment(envParameters)
    envCount += 1
  msgText = 'Imported ' + str(envCount) + ' environments'
  if (pSettings != None):
    msgText += ', and project settings'
    msgText += '.'
  return msgText

def importComponentViewFile(importFile,session_id = None):
  try:
    parser = xml.sax.make_parser()
    handler = ArchitecturalPatternContentHandler()
    parser.setContentHandler(handler)
    parser.setEntityResolver(handler)
    parser.parse(importFile)
    view = handler.view()
    return importComponentViewData(view,session_id = session_id)
  except xml.sax.SAXException as e:
    raise ARMException("Error parsing" + importFile + ": " + e.getMessage())

def importAssetsFile(importFile,session_id = None):
  try:
    parser = xml.sax.make_parser()
    handler = TemplateAssetsContentHandler()
    parser.setContentHandler(handler)
    parser.setEntityResolver(handler)
    parser.parse(importFile)
    return importAssets(handler.valueTypes(),handler.assets(),session_id)
  except xml.sax.SAXException as e:
    raise ARMException("Error parsing" + importFile + ": " + e.getMessage())

def importAssets(valueTypes,assets,session_id):
  b = Borg()
  db_proxy = b.get_dbproxy(session_id)
  vtCount = 0
  taCount = 0

  for vtParameters in valueTypes:
    vtId = db_proxy.existingObject(vtParameters.name(),vtParameters.type())
    if vtId == -1:
      db_proxy.addValueType(vtParameters)
      vtCount += 1
  for taParameters in assets:
    taId = db_proxy.existingObject(taParameters.name(),'template_asset')
    if taId == -1:
      db_proxy.addTemplateAsset(taParameters)
      taCount += 1
  return 'Imported ' + str(vtCount) + ' value types, and ' + str(taCount) + ' template assets.'

def importComponentViewData(view,session_id = None):
  b = Borg()
  db_proxy = b.get_dbproxy(session_id)
  db_proxy.addComponentView(view)
  msgStr = 'Imported architectural pattern'
  return msgStr

def importSynopsesFile(importFile,session_id = None):
  try:
    parser = xml.sax.make_parser()
    handler = SynopsesContentHandler(session_id = session_id)
    parser.setContentHandler(handler)
    parser.setEntityResolver(handler)
    parser.parse(importFile)
    charSyns = handler.characteristicSynopses()
    refSyns = handler.referenceSynopses()
    stepSyns = handler.stepSynopses()
    refConts = handler.referenceContributions()
    ucConts = handler.useCaseContributions()
    taskConts = handler.taskContributions()
    return importSynopses(charSyns,refSyns,stepSyns,refConts,ucConts,taskConts,session_id = session_id)
  except xml.sax.SAXException as e:
    raise ARMException("Error parsing" + importFile + ": " + e.getMessage())

def importSynopses(charSyns,refSyns,stepSyns,refConts,ucConts,taskConts,session_id):
  b = Borg()
  db_proxy = b.get_dbproxy(session_id)
  for cs in charSyns:
    db_proxy.addCharacteristicSynopsis(cs)
  for rs in refSyns:
    db_proxy.addReferenceSynopsis(rs)
  for ucName,envName,stepNo,synName,aType,aName in stepSyns:
    db_proxy.addStepSynopsis(ucName,envName,stepNo,synName,aType,aName)
  db_proxy.conn.commit()
  for rc in refConts:
    db_proxy.addReferenceContribution(rc)
  for uc in ucConts:
    db_proxy.addUseCaseContribution(uc)
  for tc in taskConts:
    db_proxy.addTaskContribution(tc)
  msgStr = 'Imported ' + str(len(charSyns)) + ' characteristic synopses, ' + str(len(refSyns)) + ' reference synopses, ' + str(len(stepSyns)) + ' step synopses, ' + str(len(refConts)) + ' reference contributions, ' + str(len(ucConts)) + ' use case contributions, and ' + str(len(taskConts)) + ' task contrbutions.'
  return msgStr

def importStoriesFile(importFile,session_id = None):
  try:
    parser = xml.sax.make_parser()
    handler = StoriesContentHandler(session_id = session_id)
    parser.setContentHandler(handler)
    parser.setEntityResolver(handler)
    parser.parse(importFile)
    uss = handler.userStories()
    return importStories(uss,session_id = session_id)
  except xml.sax.SAXException as e:
    raise ARMException("Error parsing" + importFile + ": " + e.getMessage())

def importStories(userStories,session_id):
  b = Borg()
  db_proxy = b.get_dbproxy(session_id)

  usCount = 0
  for userStory in userStories:
    objtId = db_proxy.existingObject(userStory.name(),'userstory')
    if objtId == -1:
      db_proxy.addUserStory(userStory)
    else:
      userStory.theId = objtId
      db_proxy.updateUserStory(userStory)
    usCount += 1

  msgStr = 'Imported ' + str(usCount) + ' user stories.'
  return msgStr


def importDomainValuesFile(importFile,session_id = None):
  try:
    parser = xml.sax.make_parser()
    handler = DomainValueContentHandler()
    parser.setContentHandler(handler)
    parser.setEntityResolver(handler)
    parser.parse(importFile)
    tvValues,rvValues,cvValues,svValues,lvValues,capValues,motValues = list(handler.values())
    return importDomainValues(tvValues,rvValues,cvValues,svValues,lvValues,capValues,motValues,session_id = session_id)
  except xml.sax.SAXException as e:
    raise ARMException("Error parsing" + importFile + ": " + e.getMessage())

def importDomainValues(tvValues,rvValues,cvValues,svValues,lvValues,capValues,motValues,session_id):
  noOfTvs = len(tvValues)
  noOfRvs = len(rvValues)
  noOfCvs = len(cvValues)
  noOfSvs = len(svValues)
  noOfLvs = len(lvValues)
  noOfCapVs = len(capValues)
  noOfMotVs = len(motValues)
 
  b = Borg()
  db_proxy = b.get_dbproxy(session_id)

  tId = 0
  if (noOfTvs > 0):
    for tvp in tvValues:
      tvp.setId(tId)
      db_proxy.updateValueType(tvp)
      tId += 1
  tId =1
  if (noOfRvs > 0):
    for rvp in rvValues:
      rvp.setId(tId)
      db_proxy.updateValueType(rvp)
      tId += 1
  tId = 0
  if (noOfCvs > 0):
    for cvp in cvValues:
      cvp.setId(tId)
      db_proxy.updateValueType(cvp)
      tId += 1
  tId = 0
  if (noOfSvs > 0):
    for svp in svValues:
      svp.setId(tId)
      db_proxy.updateValueType(svp)
      tId += 1
  tId = 0
  if (noOfLvs > 0):
    for lvp in lvValues:
      lvp.setId(tId)
      db_proxy.updateValueType(lvp)
      tId += 1
  if (noOfCapVs > 0):
    for capvp in capValues:
      db_proxy.addValueType(capvp)
  if (noOfMotVs > 0):
    for motvp in motValues:
      db_proxy.addValueType(motvp)

  msgStr = 'Imported domain values'
  return msgStr

def importProcessesFile(importFile,session_id = None):
  try:
    parser = xml.sax.make_parser()
    handler = ProcessesContentHandler()
    parser.setContentHandler(handler)
    parser.setEntityResolver(handler)
    parser.parse(importFile)
    docs = handler.internalDocuments()
    codes = handler.codes()
    memos = handler.memos()
    quotations = handler.quotations()
    codeNetworks = handler.codeNetworks()
    processes = handler.processes()
    ics = handler.impliedCharacteristics()
    intentions = handler.intentions()
    contributions = handler.contributions()
    return importProcesses(docs,codes,memos,quotations,codeNetworks,processes,ics,intentions,contributions,session_id = session_id)
  except xml.sax.SAXException as e:
    raise ARMException("Error parsing" + importFile + ": " + e.getMessage())

def importProcesses(docs,codes,memos,quotations,codeNetworks,processes,ics,intentions,contributions,session_id):
  noOfDocs = len(docs)
  noOfCodes = len(codes)
  noOfMemos = len(memos)
  noOfQuotations = len(quotations)
  noOfCNs = len(codeNetworks)
  noOfProcs = len(processes)
  noOfICs = len(ics)
  noOfIntentions = len(intentions)
  noOfContributions = len(contributions)

  b = Borg()
  db_proxy = b.get_dbproxy(session_id)

  for dp in docs:
    db_proxy.addInternalDocument(dp)

  for cp in codes:
    db_proxy.addCode(cp)

  for mp in memos:
    db_proxy.addMemo(mp)

  for q in quotations:
    db_proxy.addQuotation(q)

  # Necessary because adding document memos currently overwrites the existing memo text
  for mp in memos:
    db_proxy.updateMemo(mp)

  for cn in codeNetworks:
    personaName = cn[0]
    rtName = cn[1]
    fromCode = cn[2]
    toCode = cn[3]
    db_proxy.addCodeRelationship(personaName,fromCode,toCode,rtName)

  for p in processes:
    db_proxy.addImpliedProcess(p)

  for ic in ics:
    db_proxy.addImpliedCharacteristic(ic)

  for intention in intentions:
    db_proxy.addIntention(intention)

  for contribution in contributions:
    db_proxy.addContribution(contribution)

  msgStr = 'Imported ' + str(noOfDocs) + ' internal documents, ' + str(noOfCodes) + ' codes, ' + str(noOfMemos) + ' memos, ' + str(noOfQuotations) + ' quotations, ' + str(noOfCNs) + ' code relationships, ' + str(noOfProcs) + ' implied processes, ' + str(noOfIntentions) + ' intentions, and ' + str(noOfContributions) + ' contributions.'
  return msgStr

def importLocationsFile(importFile,session_id = None):
  try: 
    parser = xml.sax.make_parser()
    handler = LocationsContentHandler()
    parser.setContentHandler(handler)
    parser.setEntityResolver(handler)
    parser.parse(importFile)
    locations = handler.locations()
    impStr = ''
    for locs in locations:
      impStr += importLocations(locations,session_id) + '. '
    return impStr
  except xml.sax.SAXException as e:
    raise ARMException("Error parsing" + importFile + ": " + e.getMessage())
  
def importLocations(locations,session_id):
  b = Borg()
  db_proxy = b.get_dbproxy(session_id)
  msgStr = ''
  for locs in locations:
    db_proxy.addLocations(locs)
    msgStr = 'Imported ' + locs.name() + '. '
  return msgStr

def importDataflowsFile(importFile,session_id = None):
  try: 
    parser = xml.sax.make_parser()
    handler = DataflowsContentHandler()
    parser.setContentHandler(handler)
    parser.setEntityResolver(handler)
    parser.parse(importFile)
    dfs = handler.dataflows()
    tbs = handler.trustBoundaries()
    buf = importDataflows(dfs,session_id)
    buf += importTrustBoundaries(tbs,session_id)
    return buf
  except xml.sax.SAXException as e:
    raise ARMException("Error parsing" + importFile + ": " + e.getMessage())

def importDataflows(dataflows,session_id):
  b = Borg()
  db_proxy = b.get_dbproxy(session_id)
  for df in dataflows:
    db_proxy.addDataFlow(df)
  noOfDataflows = len(dataflows)
  msgStr = 'Imported ' +str( noOfDataflows) + ' dataflow'
  if (noOfDataflows != 1):
    msgStr += 's'
  msgStr += '. '
  return msgStr

def importTrustBoundaries(tbs,session_id):
  b = Borg()
  db_proxy = b.get_dbproxy(session_id)
  for tb in tbs:
    objtId = db_proxy.existingObject(tb.name(),'trust_boundary')
    if objtId == -1:
      db_proxy.addTrustBoundary(tb)
    else:
      tb.setId(objtId)
      db_proxy.updateTrustBoundary(tb)
  noOfTrustBoundaries = len(tbs)
  msgStr = 'Imported ' +str( noOfTrustBoundaries) + ' trust boundar'
  if (noOfTrustBoundaries != 1):
    msgStr += 'ies'
  else:
    msgStr += 'y'
  msgStr += '.'
  return msgStr

def importModelFile(importFile,isOverwrite = 1,session_id = None):
  try:
    b = Borg()
    db_proxy = b.dbProxy
    if (session_id != None):
      db_proxy = b.get_dbproxy(session_id)

    modelTxt = ''
    if isOverwrite == 1:
      db_proxy.clearDatabase(session_id)
      modelTxt += importTVTypeFile(importFile,isOverwrite,session_id = session_id) + '  '
    modelTxt += importDomainValuesFile(importFile,session_id) + ' '
    modelTxt += importProjectFile(importFile,session_id) + ' '
    modelTxt += importRiskAnalysisFile(importFile,session_id) + ' '
    modelTxt += importUsabilityFile(importFile,session_id) + ' '
    modelTxt += importRequirementsFile(importFile,session_id) + ' '
    modelTxt += importAssociationsFile(importFile,session_id) + ' '
    modelTxt += importSynopsesFile(importFile,session_id) + ' '
    modelTxt += importMisusabilityFile(importFile,session_id) + ' '
    modelTxt += importDataflowsFile(importFile,session_id) + ' '
    modelTxt += importLocationsFile(importFile,session_id) + ' '
    modelTxt += importStoriesFile(importFile,session_id)
    return modelTxt
  except xml.sax.SAXException as e:
    raise ARMException("Error parsing" + importFile + ": " + e.getMessage())
  except Exception as e:
    raise ARMException("Error importing " + importFile + ": " + format(e))

def importAttackTreeString(buf,session_id = None):
  try:
    b = Borg()
    db_proxy = b.get_dbproxy(session_id)
    modelTxt = ''
    modelTxt += importRequirementsString(buf,session_id) + ' '
    modelTxt += importAssociationsString(buf,session_id) + ' '
    return modelTxt
  except xml.sax.SAXException as e:
    raise ARMException("Error parsing" + importFile + ": " + e.getMessage())

def importDiagramsNetFile(importFile,modelType):
  try:
    parser = xml.sax.make_parser()
    handler = DiagramsNetContentHandler(modelType)
    parser.setContentHandler(handler)
    parser.setEntityResolver(handler)
    parser.parse(importFile)
    if (modelType == 'dataflow'):
      return (handler.objects(),handler.flows(), handler.trustBoundaries())
    else:
      return (handler.objects(),handler.associations())
  except xml.sax.SAXException as e:
    raise ARMException("Error parsing" + importFile + ": " + e.getMessage())

def importDiagramsNetDFD(importFile,envName,session_id):
  objts, flows, tbs = importDiagramsNetFile(importFile,'dataflow')
  b = Borg()
  db_proxy = b.get_dbproxy(session_id)

  newAssetCount = 0
  newUCCount = 0
  for objt in objts:
    objtType = objt['type']
    assetType = 'Information'
    if (objtType == 'process'):
      objtType = 'usecase'
    elif (objtType == 'entity'):
      objtType = 'asset'
      assetType = 'Systems'
    else:
      objtType = 'asset'

    if (db_proxy.existingObject(objt['name'],objtType) == -1):
      cairis.core.DefaultParametersFactory.build(objt['name'],envName,objtType,session_id,assetType)
      if (objtType == 'usecase'):
        newUCCount += 1
      else:
        newAssetCount += 1

  flowCount = 0
  for f in flows:
    dfName = f['name']
    fromName = f['from_name']
    fromType = f['from_type']
    toName = f['to_name']
    toType = f['to_type']
    try:
       db_proxy.checkDataFlowExists(dfName, fromType, fromName, toType, toName, envName)
       dfAssets = f['assets']
       for dfAsset in dfAssets:
         if (db_proxy.existingObject(dfAsset,'asset') == -1):
           cairis.core.DefaultParametersFactory.build(dfAsset,envName,'asset',session_id,'Information')
           newAssetCount += 1
       db_proxy.addDataFlow(DataFlowParameters(dfName,'Information',envName,fromName,fromType,toName,toType,dfAssets))
       flowCount += 1
    except DatabaseProxyException as ex:
      if str(ex.value).find('already exists') == -1:
        raise ARMException(str(ex.value))
    except ARMException as ex:
      if str(ex.value).find('already exists') == -1:
        raise ARMException(str(ex.value))

  tbCount = 0
  for tb in tbs:
    tbName = tb['name']
    if (db_proxy.existingObject(tbName,'trust_boundary') == -1):
      tbComps = []
      for tbComp in tb['components']:
        tbComps.append((tbComp['type'],tbComp['name']))
      db_proxy.addTrustBoundary(TrustBoundary(-1,tbName,'General','To be defined',{envName : tbComps},{envName : 'None'},[])) 
      tbCount += 1

  msgStr = 'Imported ' + str(newAssetCount) + ' asset'
  if (newAssetCount != 1): msgStr += 's'
  msgStr += ', ' + str(newUCCount) + ' use case'
  if (newUCCount != 1): msgStr += 's'
  msgStr += ', ' + str(flowCount) + ' data flow'
  if (flowCount != 1): msgStr += 's'
  msgStr += ', and ' + str(tbCount) + ' trust boundar'
  if (tbCount != 1): 
    msgStr += 'ies'
  else:
    msgStr += 'y'
  msgStr += '.'
  return msgStr

def importDiagramsNetAssetModel(importFile,envName,session_id):
  objts, assocs = importDiagramsNetFile(importFile,'asset')
  b = Borg()
  db_proxy = b.get_dbproxy(session_id)

  newAssetCount = 0
  for asset in objts:
    assetName = asset['name']
    assetType = asset['type'].capitalize()

    if (db_proxy.existingObject(assetName,'asset') == -1):
      db_proxy.addAsset(AssetParameters(assetName,asset['short_code'],asset['description'],asset['significance'],assetType,0,'',[],[],[AssetEnvironmentProperties(envName,asset['properties'],asset['rationale'])]))
      newAssetCount += 1

  newAssocCount = 0
  for assoc in assocs:
    headAsset = assoc['head']
    tailAsset = assoc['tail']
    try:
      db_proxy.checkAssetAssociation(envName,headAsset,tailAsset)
      db_proxy.addClassAssociation(ClassAssociationParameters(envName,headAsset,'asset',assoc['headNav'],assoc['headType'],'*','','','*',assoc['tailType'],assoc['tailNav'],'asset',tailAsset,'To be defined'))
      newAssocCount += 1
    except DatabaseProxyException as ex:
      if str(ex.value).find('already exists') == -1:
        raise ARMException(str(ex.value))
  
  msgStr = 'Imported ' + str(newAssetCount) + ' asset'
  if (newAssetCount != 1): msgStr += 's'
  msgStr += ', and ' + str(newAssocCount) + ' asset association'
  if (newAssocCount != 1): msgStr += 's'
  msgStr += '.'
  return msgStr

def importUserGoalWorkbook(wbFile,session_id):
  charSyns = []
  refSyns = []
  refConts = []

  wb = load_workbook(filename = wbFile,data_only=True)
  ugSheet = wb.worksheets[0]
  for row in ugSheet.iter_rows(min_row=2):
    refName = row[0].value
    refDesc = row[1].value
    pName = row[2].value
    pdRef = row[3].value
    elType = row[4].value
    ugName = row[5].value
    if (ugName != '' and ugName != None and ugName != 0):
      initSat = row[6].value
      if (pdRef == 'persona'):
        charSyns.append(ReferenceSynopsis(-1,refName,ugName,elType,'persona',pName,'persona_characteristic',initSat))
      else:
        refSyns.append(ReferenceSynopsis(-1,refName,ugName,elType,'persona',pName,'document_reference',initSat))

  contSheet = wb.worksheets[1]
  for row in contSheet.iter_rows(min_row=2):
    srcName = row[0].value
    destName = row[1].value
    meName = row[2].value
    contName = row[3].value
    if (srcName != '' and srcName != None and srcName != 0 and destName != '' and destName != None and destName != 0):
      refConts.append(ReferenceContribution(srcName,destName,meName,contName))

  b = Borg()
  db_proxy = b.get_dbproxy(session_id)
  for cs in charSyns:
    db_proxy.addCharacteristicSynopsis(cs)
  for rs in refSyns:
    db_proxy.addReferenceSynopsis(rs)
  db_proxy.conn.commit()
  for rc in refConts:
    db_proxy.addReferenceContribution(rc)

  msgStr = 'Imported ' + str(len(charSyns)) + ' characteristic synopses, ' + str(len(refSyns)) + ' reference synopses, and ' + str(len(refConts)) + ' reference contributions.'
  return msgStr

def importPersonaCharacteristicsWorkbook(wbFile,session_id):
  eds = []
  drs = []
  pcs = []

  wb = load_workbook(filename = wbFile,data_only=True)
  edSheet = wb.worksheets[0]
  edNames = set([])
  for row in edSheet.iter_rows(min_row=2):
    edName = row[0].value
    edAuth = row[1].value
    edVer = row[2].value
    edPD = str(row[3].value)
    edDesc = row[4].value
    if (edName != '' and edName != None):
      edNames.add(edName)
      if (edAuth == '' or edAuth == None):
        edAuth = 'Unknown'
      if (edVer == '' or edVer == None):
        edVer = '1'
      if (edPD == '' or edPD == None):
        edPD = 'Unknown'
      if (edDesc == '' or edDesc == None):
        edDesc = 'Unknown'
      eds.append(ExternalDocumentParameters(edName,edVer,edPD,edAuth,edDesc))

  b = Borg()
  db_proxy = b.get_dbproxy(session_id)

  drNames = set([])
  drSheet = wb.worksheets[1]
  for row in drSheet.iter_rows(min_row=2):
    drName = row[0].value
    drDoc = row[1].value
    drCont = row[2].value
    drExc = row[3].value

    if (drName != '' and drName != None):
      drNames.add(drName)
      if (drDoc not in edNames):
        if (db_proxy.existingObject(drDoc,'external_document') == -1):
          raise ARMException("Cannot import document reference " + drName + ". " + drDoc + " not an external document.")
      if (drCont == '' or drCont == None):
        drCont = 'Unknown'
      if (drExc == '' or drExc == None):
        drExc = 'Unknown'
      drs.append(DocumentReferenceParameters(drName,drDoc,drCont,drExc))

  pcSheet = wb.worksheets[2]
  for row in pcSheet.iter_rows(min_row=2):
    pcName = row[0].value
    pName = row[1].value.strip()
    bvName = row[2].value.strip()
    modQual = row[3].value.strip()
    grounds  = row[4].value
    if (grounds == None):
      grounds = []
    else:
      grounds = list(map(lambda x: (x.strip(),'','document'),grounds.strip().split(',')))
      if (len(grounds) == 1 and grounds[0][0] == ''):
        grounds = []

    warrant  = row[5].value
    if (warrant == None):
      warrant = []
    else:
      warrant = list(map(lambda x: (x.strip(),'','document'),warrant.strip().split(',')))
      if (len(warrant) == 1 and warrant[0][0] == ''):
        warrant = []

    rebuttal = row[6].value
    if (rebuttal == None):
      rebuttal = []
    else:
      rebuttal = list(map(lambda x: (x.strip(),'','document'),rebuttal.strip().split(',')))
      if (len(rebuttal) == 1 and rebuttal[0][0] == ''):
        rebuttal = []

    if (pcName != ''):
      if (db_proxy.existingObject(pName,'persona') == -1):
        raise ARMException("Cannot import persona characteristic " + pcName + ". Persona " + pName + " does not exist.")
      if bvName not in ['Activities','Attitudes','Aptitudes','Motivations','Skills','Environment Narrative','Intrinsic','Contextual']:
        raise ARMException("Cannot import persona characteristic " + pcName + ". " + bvName + " is an invalid behavioural variable name.")

      if (modQual == '' or modQual == None):
        modQual = 'Possibly'

      for g in grounds:
        if ((g[0] not in drNames) and (db_proxy.existingObject(g[0],'document_reference') == -1)):
          raise ARMException("Cannot import persona characteristic " + pcName + ". Document reference corresponding with grounds " + g[0] + " does not exist.")
      for w in warrant:
        if ((w[0] not in drNames) and (db_proxy.existingObject(w[0],'document_reference') == -1)):
          raise ARMException("Cannot import persona characteristic " + pcName + ". Document reference corresponding with warrant " + w[0] + " does not exist.")
      for r in rebuttal:
        if ((r[0] not in drNames) and (db_proxy.existingObject(r[0],'document_reference') == -1)):
          raise ARMException("Cannot import persona characteristic " + pcName + ". Document reference corresponding with rebuttal " + r[0] + " does not exist.")

      pcs.append(PersonaCharacteristicParameters(pName,modQual,bvName,pcName,grounds,warrant,[],rebuttal))
      
  for edParameters in eds:
    objtId = db_proxy.existingObject(edParameters.name(),'external_document')
    if objtId == -1:
      db_proxy.addExternalDocument(edParameters)
    else:
      edParameters.setId(objtId)
      db_proxy.updateExternalDocument(edParameters)

  for drParameters in drs:
    objtId = db_proxy.existingObject(drParameters.name(),'document_reference')
    if objtId == -1:
      db_proxy.addDocumentReference(drParameters)
    else:
      drParameters.setId(objtId)
      db_proxy.updateDocumentReference(drParameters)

  for pcParameters in pcs:
    objtId = db_proxy.existingObject(pcParameters.characteristic(),'persona_characteristic')
    if objtId == -1:
      db_proxy.addPersonaCharacteristic(pcParameters)
    else:
      pcParameters.setId(objtId)
      db_proxy.updatePersonaCharacteristic(pcParameters)

  msgStr = 'Imported ' + str(len(eds)) + ' external documents, ' + str(len(drs)) + ' document references, and ' + str(len(pcs)) + ' persona characteristics.'
  return msgStr
